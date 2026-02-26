import openpyxl
import os

wb = openpyxl.load_workbook('Equipment Inspections SiMSET.xlsx', data_only=True)

# 1. Parse Courses
ws_work = wb['WorkList_Data']
courses_sql = []
courses_sql.append("INSERT INTO courses (work_id, work_type, name) VALUES")
courses_values = []
seen_work_id = set()

for row in ws_work.iter_rows(min_row=2, values_only=True):
    work_id = row[0]
    work_type = row[1]
    name = row[2]
    if not work_id or work_id in seen_work_id:
        continue
    seen_work_id.add(work_id)
    # escape single quotes
    w_type = str(work_type).replace("'", "''") if work_type else ''
    w_name = str(name).replace("'", "''") if name else ''
    courses_values.append(f"('{work_id}', '{w_type}', '{w_name}')")

courses_sql.append(",\n".join(courses_values) + "\nON CONFLICT (work_id) DO NOTHING;")

with open('insert_courses.sql', 'w', encoding='utf-8') as f:
    f.write("\n".join(courses_sql))


# 2. Parse Manikins
ws_manikins = wb['รายการหุ่น']
manikins_sql = []
manikins_sql.append("INSERT INTO manikins (sap_id, team_code, asset_code, asset_name, status, is_active) VALUES")
manikins_values = []
seen_sap = set()

for i, row in enumerate(ws_manikins.iter_rows(min_row=2, values_only=True)):
    team_code = str(row[0]).replace("'", "''") if row[0] else None
    if not team_code:
        continue
        
    raw_sap = row[1]
    if raw_sap:
        # Convert float like 131200000516.0 to string '131200000516'
        try:
            sap_id = str(int(float(raw_sap)))
        except ValueError:
            sap_id = str(raw_sap).strip()
    else:
        # Fallback sap_id to team_code if empty
        sap_id = f"NO-SAP-{team_code}"
        
    # Skip duplicates just in case
    if sap_id in seen_sap:
        continue
    seen_sap.add(sap_id)

    asset_code = str(row[2]).replace("'", "''") if row[2] else ''
    asset_name = str(row[3]).replace("'", "''") if row[3] else 'Unknown'
    
    thai_status = row[5]
    status_enum = 'ready'
    is_active = True
    
    if thai_status == 'พร้อมใช้งาน':
        status_enum = 'ready'
    elif thai_status == 'อยู่ระหว่างการซ่อม':
        status_enum = 'maintenance'
    elif thai_status == 'ไม่พร้อมใช้งาน':
        status_enum = 'broken'
    elif thai_status == 'โอนย้าย':
        # Don't have a status enum for transferred, but we set is_active=False
        status_enum = 'ready' 
        is_active = False

    manikins_values.append(f"('{sap_id}', '{team_code}', '{asset_code}', '{asset_name}', '{status_enum}', {str(is_active).lower()})")

manikins_sql.append(",\n".join(manikins_values) + "\nON CONFLICT (sap_id) DO UPDATE SET is_active = EXCLUDED.is_active;")

with open('insert_manikins.sql', 'w', encoding='utf-8') as f:
    f.write("\n".join(manikins_sql))

print(f"Generated insert_courses.sql with {len(courses_values)} courses")
print(f"Generated insert_manikins.sql with {len(manikins_values)} manikins")
