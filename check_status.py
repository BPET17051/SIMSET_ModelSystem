import openpyxl
wb = openpyxl.load_workbook('Equipment Inspections SiMSET.xlsx')
ws = wb['รายการหุ่น']

col_headers = [cell.value for cell in ws[1]]
print('Columns:', col_headers[:8])

statuses = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    status = row[5]  # สถานะหุ่น = column index 5
    if status:
        statuses[status] = statuses.get(status, 0) + 1

print('\n=== สถานะหุ่น (unique values) ===')
for s, count in sorted(statuses.items(), key=lambda x: -x[1]):
    print(f'  "{s}"  -> {count} ตัว')
print(f'\nTotal rows with status: {sum(statuses.values())}')
