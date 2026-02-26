import openpyxl
import json
import urllib.request

SUPABASE_URL = 'https://ifogcvymwhcfbfjzhwsl.supabase.co/rest/v1/'
API_KEY = 'sb_publishable_DZyIDHVZ-kfD1o3baz0qmw_tTyRCJG8'

headers = {
    'apikey': API_KEY,
    'Authorization': f'Bearer {API_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'resolution=ignore-duplicates'
}

wb = openpyxl.load_workbook('Equipment Inspections SiMSET.xlsx', data_only=True)

# 1. Import Courses
print("Importing courses...")
ws_work = wb['WorkList_Data']
courses = []
seen_work_id = set()

for row in ws_work.iter_rows(min_row=2, values_only=True):
    work_id = row[0]
    work_type = row[1]
    name = row[2]
    if not work_id or work_id in seen_work_id:
        continue
    seen_work_id.add(work_id)
    courses.append({
        "work_id": str(work_id),
        "work_type": str(work_type) if work_type else '',
        "name": str(name) if name else ''
    })

req = urllib.request.Request(SUPABASE_URL + 'courses', data=json.dumps(courses).encode('utf-8'), headers=headers, method='POST')
try:
    with urllib.request.urlopen(req) as response:
        print(f"Courses imported! Status: {response.status}")
except Exception as e:
    print(f"Error importing courses: {e}")

# 2. Import Manikins
print("Importing manikins...")
ws_manikins = wb['รายการหุ่น']
manikins = []
seen_sap = set()

for row in ws_manikins.iter_rows(min_row=2, values_only=True):
    team_code = str(row[0]) if row[0] else None
    if not team_code:
        continue
        
    raw_sap = row[1]
    if raw_sap:
        try:
            sap_id = str(int(float(raw_sap)))
        except ValueError:
            sap_id = str(raw_sap).strip()
    else:
        sap_id = f"NO-SAP-{team_code}"
        
    if sap_id in seen_sap:
        continue
    seen_sap.add(sap_id)

    asset_code = str(row[2]) if row[2] else ''
    asset_name = str(row[3]) if row[3] else 'Unknown'
    
    thai_status = row[5]
    status_enum = 'ready'
    is_active = True
    
    if thai_status == 'พร้อมใช้งาน':
        status_enum = 'ready'
    elif thai_status == 'อยู่ระหว่างการซ่อม':
        status_enum = 'maintenance'
    elif thai_status == 'ไม่พร้อมใช้งาน':
        status_enum = 'broken'
    elif thai_status == 'โอนย้าย':
        status_enum = 'ready' 
        is_active = False

    manikins.append({
        "sap_id": sap_id,
        "team_code": team_code,
        "asset_code": asset_code,
        "asset_name": asset_name,
        "status": status_enum,
        "is_active": is_active,
        "needs_review": False
    })

req = urllib.request.Request(SUPABASE_URL + 'manikins', data=json.dumps(manikins).encode('utf-8'), headers=headers, method='POST')
try:
    with urllib.request.urlopen(req) as response:
        print(f"Manikins imported! Status: {response.status}")
except urllib.error.HTTPError as e:
    err = e.read().decode('utf-8')
    print(f"HTTPError importing manikins: {e.code} - {err}")
except Exception as e:
    print(f"Error importing manikins: {e}")
